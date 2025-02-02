import pandas as pd
import os
import datetime as dt
import time

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Color, Alignment
from openpyxl.formula.translate import Translator
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

import docx
from docx import Document

import tkinter as tk
from tkinter import ttk

# import pyautogui
# import PySimpleGUI as sg

# Atribuindo caminho dos diretórios
diretorio = os.getcwd()

path_bases_auxiliares = diretorio + '/bases_auxiliares/'
path_fornecedores = diretorio + '/fornecedores/'
path_catalogo = diretorio + '/catalogo/'
path_template = diretorio + '/templates/'
path_pedidos_flores = diretorio + '/pedidos/'

# Carregando template
wb = load_workbook(filename = path_template + 'Catalogo_flores.xlsx')
ws = wb.active

# Padronizando estilos da planilha
fonte = Font(name='Calibri', size=10, color="000000", bold = False)
borda = Side(border_style="thin", color="000000")
cor_fundo = PatternFill("solid", fgColor="00FFFFFF")
cor_rodape = PatternFill("solid",fgColor="F4B082")

# Inserindo linhas no template
ws[f"B3"].value = dt.datetime.now().strftime('%d-%m-%Y')
       

#Separando tipos de arquivos
arqs_fornecedores_xlsx = pd.DataFrame()
arqs_fornecedores_jpg = pd.DataFrame()

# Carregando bases auxiliares
base_de_cores = pd.DataFrame()
base_de_cores = pd.read_excel(path_bases_auxiliares + 'base_de_cores.xlsx', header=0)

# Lendo pastas de fornecedores
diretorio_fornecedores = os.listdir(path_fornecedores)


# ----- Criando função para atualização do texto da barra de progresso da interface -----
def atualiza_label_progresso(etapa):
    
    if etapa == 1:
        texto_status['text'] = "Consolidando planilhas, por favor aguarde..."
    elif etapa == 2:
        texto_status['text'] = "Consolidação concluída!"
    elif etapa == 3:
        texto_status['text'] = "Realizando organização dos arquivos..."
    elif etapa == 4:
        texto_status['text'] = "Organização concluída!"

    return texto_status['text']

# ----- Criando função para atualização da barra de progresso da interface -----
def atualiza_barra_progresso(valor):
    
    valor_atual = barra_progresso['value']

    if valor_atual < 101:
        barra_progresso['value'] = valor_atual + valor
        janela.update_idletasks()
    
# -------------------------------------------------------------------------------

# ----- Iniciando processo de importação das bases de fornecedores -----
def consolida_fornecedor():

# Lendo bases de fornecedor 1
    arqs_fornecedores = os.listdir(path_fornecedores + 'fornecedor 1/')
    
    for arqs in range(len(arqs_fornecedores)):
        if len(arqs_fornecedores) > 0:

            #Separando arquivos .xlsx
            if arqs_fornecedores[arqs].find(".xlsx") != -1:
                
                # Lendo sheets da planilha
                xlsx_sheets_temp = pd.ExcelFile(path_fornecedores + 'fornecedor 1/' + arqs_fornecedores[arqs]).sheet_names
                            
                for sheets in range(len(xlsx_sheets_temp)):
                    
                    xlsx_temp = pd.DataFrame()
                    xlsx_temp = pd.read_excel(path_fornecedores + 'fornecedor 1/' + arqs_fornecedores[arqs], sheet_name = xlsx_sheets_temp[sheets], header=None)
                    xlsx_temp.columns = ['Flores','Valor']
                    xlsx_temp = xlsx_temp.dropna()
                    xlsx_temp = xlsx_temp.join(base_de_cores.set_index('Flores'), how='left', on='Flores')

                    if sheets == 0:

                        # Verificando total de linhas da base
                        total_linhas = len(xlsx_temp)
                        print(f"sheet 0 = {total_linhas}")

                        for linha in range(total_linhas):
                        
                            ws[f"A{linha + 6}"].value = xlsx_temp.iat[linha,0]

                            ws[f"B{linha + 6}"].value = xlsx_temp.iat[linha,3]

                            ws[f"C{linha + 6}"].value = xlsx_temp.iat[linha,4]
                            
                            # ws[f"D{linha + 6}"].value = xlsx_sheets_temp[sheets]
                            ws[f"D{linha + 6}"].value = xlsx_temp.iat[linha,2]
                            
                            ws[f"E{linha + 6}"].value = "Fornecedor 1"

                            ws[f"G{linha + 6}"].value = xlsx_temp.iat[linha,1]
                            ws[f"G{linha + 6}"].number_format = 'R$ #,##0.00'

                            ws[f"H{linha + 6}"].value = f"=F{linha + 6}*G{linha + 6}"
                            ws[f"H{linha + 6}"].number_format = 'R$ #,##0.00' 

                            ws[f"A{linha + 6}"].font = fonte
                            ws[f"B{linha + 6}"].font = fonte
                            ws[f"C{linha + 6}"].font = fonte
                            ws[f"D{linha + 6}"].font = fonte
                            ws[f"E{linha + 6}"].font = fonte
                            ws[f"F{linha + 6}"].font = fonte
                            ws[f"G{linha + 6}"].font = fonte
                            ws[f"H{linha + 6}"].font = fonte
                            
                            ws[f"A{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"B{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"C{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"D{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"E{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"F{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"G{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"H{linha + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)

                            ws[f"A{linha + 6}"].fill = cor_fundo
                            ws[f"B{linha + 6}"].fill = cor_fundo
                            ws[f"C{linha + 6}"].fill = cor_fundo
                            ws[f"D{linha + 6}"].fill = cor_fundo
                            ws[f"E{linha + 6}"].fill = cor_fundo
                            ws[f"F{linha + 6}"].fill = cor_fundo
                            ws[f"G{linha + 6}"].fill = cor_fundo
                            ws[f"H{linha + 6}"].fill = cor_fundo

                        total_linhas_sheet_anterior = total_linhas
                        print(f"total_linhas_sheet_anterior = {total_linhas_sheet_anterior}")
                    
                    else:

                        total_linhas = len(xlsx_temp)
                        print(f"sheet {sheets} = {total_linhas}")
                        print(f"total_linhas_sheet_anterior = {total_linhas_sheet_anterior}")

                        for linha in range(total_linhas):
            
                            ws[f"A{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha,0]
                            
                            ws[f"B{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha,3]

                            ws[f"C{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha,4]
                            
                            # ws[f"D{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_sheets_temp[sheets]
                            ws[f"D{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha,2]

                            ws[f"E{linha + total_linhas_sheet_anterior + 6}"].value = "Fornecedor 1"
                            
                            ws[f"G{linha + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha,1]
                            ws[f"G{linha + total_linhas_sheet_anterior + 6}"].number_format = 'R$ #,##0.00'

                            ws[f"H{linha + total_linhas_sheet_anterior + 6}"].value = f"=F{linha + total_linhas_sheet_anterior + 6}*G{linha + total_linhas_sheet_anterior + 6}"
                            ws[f"H{linha + total_linhas_sheet_anterior + 6}"].number_format = 'R$ #,##0.00' 

                            ws[f"A{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"B{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"C{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"D{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"E{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"F{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"G{linha + total_linhas_sheet_anterior + 6}"].font = fonte
                            ws[f"H{linha + total_linhas_sheet_anterior + 6}"].font = fonte

                            ws[f"A{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"B{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"C{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"D{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"E{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"F{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"G{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                            ws[f"H{linha + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)

                            ws[f"A{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"B{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"C{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"D{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"E{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"F{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"G{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                            ws[f"H{linha + total_linhas_sheet_anterior + 6}"].fill = cor_fundo

                        total_linhas_sheet_anterior = total_linhas_sheet_anterior + total_linhas

    # Lendo bases de fornecedor 2
    arqs_fornecedores = os.listdir(path_fornecedores + 'fornecedor 2/')
        
    for arqs in range(len(arqs_fornecedores)):
        if len(arqs_fornecedores) > 0:

            #Separando arquivos .xlsx
            if arqs_fornecedores[arqs].find(".xlsx") != -1:
                # Lendo sheets da planilha
                xlsx_sheets_temp = pd.ExcelFile(path_fornecedores + 'fornecedor 2/' + arqs_fornecedores[arqs]).sheet_names
                            
                for sheets in range(len(xlsx_sheets_temp)):
                    
                    xlsx_temp = pd.DataFrame()
                    xlsx_temp = pd.read_excel(path_fornecedores + 'fornecedor 2/' + arqs_fornecedores[arqs], sheet_name = xlsx_sheets_temp[sheets], header=None, skiprows=3)
                    xlsx_temp.columns = ['Flores','Valor']
                    xlsx_temp = xlsx_temp.dropna()
                    xlsx_temp = xlsx_temp.join(base_de_cores.set_index('Flores'), how='left', on='Flores')
                    
                    # Verificando total de linhas da base
                    total_linhas = len(xlsx_temp)
                    print(f"sheet {sheets} = {total_linhas}")
                    print(f"total_linhas_sheet_anterior = {total_linhas_sheet_anterior}")
                    
                    for linha_fornecedor_2 in range(total_linhas):
                        
                        ws[f"A{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha_fornecedor_2,0]
                        
                        ws[f"B{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha_fornecedor_2,3]

                        ws[f"C{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha_fornecedor_2,4]
                        
                        # ws[f"D{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = "N/A"
                        ws[f"D{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha_fornecedor_2,2]

                        ws[f"E{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = "Fornecedor 2"
                        
                        ws[f"G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = xlsx_temp.iat[linha_fornecedor_2,1]
                        ws[f"G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].number_format = 'R$ #,##0.00'

                        ws[f"H{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].value = f"=F{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}*G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"
                        ws[f"H{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].number_format = 'R$ #,##0.00' 

                        ws[f"A{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"B{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"C{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"D{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"E{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"F{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte
                        ws[f"H{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].font = fonte

                        ws[f"A{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"B{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"C{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"D{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"E{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"F{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
                        ws[f"H{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)

                        ws[f"A{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"B{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"C{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"D{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"E{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"F{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"G{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo
                        ws[f"H{linha_fornecedor_2 + total_linhas_sheet_anterior + 6}"].fill = cor_fundo

                    total_linhas_sheet_anterior = total_linhas_sheet_anterior + total_linhas
                        
    # Editando cor das células do rodapé da tabela
    ws[f"A{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"B{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"C{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"D{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"E{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"F{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"G{total_linhas_sheet_anterior + 6}"].fill = cor_rodape
    ws[f"H{total_linhas_sheet_anterior + 6}"].fill = cor_rodape

    # Inserindo bordas nas células de rodapé da tabela
    ws[f"A{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"B{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"C{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"D{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"E{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"F{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"G{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)
    ws[f"H{total_linhas_sheet_anterior + 6}"].border = Border(top=borda, left=borda, right=borda, bottom=borda)



    ws.sheet_view.showGridLines = False
    wb.save(path_catalogo + 'Catalogo Flores - ' + dt.datetime.now().strftime('%d-%m-%Y') + '.xlsx')
    wb.close()
    
    transferir_arquivos()


# ----- Transferindo versões de catálogos antigos para o backlog -----
def transferir_arquivos():
    # Criando diretorio de backlog caso não exista
    if not (os.path.isdir(path_catalogo + '_old')):
        os.mkdir(path_catalogo + '_old')

    arquivos_catalogos_old = os.listdir(path_catalogo)

    for arqs_old in range(len(arquivos_catalogos_old)):
        
        if (arquivos_catalogos_old[arqs_old] != 'Catalogo Flores - ' + dt.datetime.now().strftime('%d-%m-%Y') + '.xlsx') and (arquivos_catalogos_old[arqs_old].find('.xlsx') != -1):
            os.replace(path_catalogo + arquivos_catalogos_old[arqs_old], path_catalogo + '_old/' + arquivos_catalogos_old[arqs_old])
    
    # Atualizando barra de progresso com os status da aplicação
    atualiza_label_progresso(1)
    atualiza_barra_progresso(15)
    time.sleep(3)

    atualiza_label_progresso(2)
    atualiza_barra_progresso(35)
    time.sleep(3)

    atualiza_label_progresso(3)
    atualiza_barra_progresso(15)
    time.sleep(3)

    atualiza_label_progresso(4)
    atualiza_barra_progresso(35)

    # Removendo botão executar e renomeando botão cancelar para encerrar a aplicação
    botao_executar.grid_remove()
    botao_cancelar['text']="Encerrar"

# --------------------------------------------------------------------

# ----- Criando interface para execução da automação -----

janela = tk.Tk()
janela.title("Jolly Design Floral - Automação")
# janela.geometry("400x200")

frame_geral = ttk.Frame(janela)
frame_geral.grid(column=0, row=0, sticky=(tk.N,tk.W,tk.E,tk.S))
janela.columnconfigure(0,weight=1)
janela.rowconfigure(0,weight=1)

frame_labels = ttk.Frame(frame_geral)
frame_labels.grid(column=0, row=0, sticky=(tk.W,tk.E))

frame_barra = ttk.Frame(frame_geral)
frame_barra.grid(column=0, row=1, sticky=(tk.W,tk.E))

frame_botoes = ttk.Frame(frame_geral)
frame_botoes.grid(column=0, row=2, sticky=(tk.W,tk.E))

# Configurando texto da interface
texto_orientacao_1 = ttk.Label(frame_labels, text="Consolidação das planilhas de fornecedores.")
texto_orientacao_1.grid(column=0, row=0, sticky=(tk.W,tk.E))

texto_orientacao_2 = ttk.Label(frame_labels, text="Clique em executar para iniciar!")
texto_orientacao_2.grid(column=0, row=1, sticky=(tk.W,tk.E))


# Configurando barra de progresso
barra_progresso = ttk.Progressbar(frame_barra, orient='horizontal', length=300, mode='determinate')
barra_progresso.grid(column=0, row=0, sticky=(tk.W,tk.E))

texto_status = ttk.Label(frame_barra, text="")
texto_status.grid(column=0, row=1, sticky=(tk.W,tk.E))


# Configurando botões
botao_executar = ttk.Button(frame_botoes, text="Executar", command= consolida_fornecedor)
botao_executar.grid(column=0, row=1, sticky=(tk.W,tk.E))

botao_cancelar = ttk.Button(frame_botoes, text="Cancelar", command= lambda: janela.quit())
botao_cancelar.grid(column=1, row=1, sticky=(tk.W,tk.E))

# Ajustando espaçamento dos frames
for widget in frame_geral.winfo_children():
    widget.grid(padx=10, pady=5)

for widget in frame_labels.winfo_children():
    widget.grid(padx=5, pady=5)

for widget in frame_barra.winfo_children():
    widget.grid(padx=5, pady=2)

for widget in frame_botoes.winfo_children():
    widget.grid(padx=5, pady=2)

janela.mainloop()

# -----------------------------------------------------------

